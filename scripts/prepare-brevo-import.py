#!/usr/bin/env python3
"""
Prépare l'import des adhérents/donateurs ADESZ dans Brevo.

Lit les fichiers Excel 2025 et 2026, fusionne les données,
et génère des CSV prêts pour l'import Brevo.

Fichiers générés :
  - brevo-import-tous.csv         → importer dans liste "Tous les contacts ADESZ" (5)
  - brevo-import-adherents.csv    → importer dans liste "Adhérents ADESZ" (3)
  - brevo-import-donateurs.csv    → importer dans liste "Donateurs ADESZ" (4)
  - contacts-sans-email.csv       → contacts sans email (à compléter manuellement)

Usage: python3 scripts/prepare-brevo-import.py
"""

import csv
import os
import re
import sys
from collections import OrderedDict

import openpyxl

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "brevo-import")

FILE_2025 = os.path.join(PROJECT_ROOT, "Association adesz.Liste des dons 2025.xlsx")
FILE_2026 = os.path.join(PROJECT_ROOT, "Association adesz.Liste des dons 2026.xlsx")

# Brevo attribute names (must match what was created in setup-brevo.php)
BREVO_COLUMNS = [
    "EMAIL", "PRENOM", "NOM", "ADRESSE", "CODE_POSTAL",
    "COMMUNE", "TYPE", "MONTANT",
]


def clean_prenom(prenom: str) -> str:
    """Remove trailing numbers/notes from prenom field."""
    if not prenom:
        return ""
    # Strip trailing digits and whitespace (e.g., "Jean 3" -> "Jean")
    cleaned = re.sub(r"\s+\d+\s*$", "", prenom.strip())
    return cleaned.strip()


def clean_email(email: str) -> str | None:
    """Extract a valid email from the field. Returns first email if multiple."""
    if not email:
        return None
    email = str(email).strip()
    if not email or "@" not in email:
        return None
    # Some fields have multiple emails separated by space
    parts = email.split()
    # Take the first valid-looking email
    for part in parts:
        if "@" in part and "." in part:
            return part.lower().strip()
    return None


def parse_address(address: str) -> dict:
    """Parse a French address string to extract components."""
    result = {"full": "", "code_postal": "", "commune": ""}
    if not address:
        return result

    address = str(address).strip()
    result["full"] = address

    # Try to extract postal code (5 digits) and commune
    match = re.search(r"(\d{5})\s+(.+?)$", address)
    if match:
        result["code_postal"] = match.group(1)
        result["commune"] = match.group(2).strip()

    return result


def parse_total(value) -> float:
    """Parse total donation amount, handling formulas like '100+100+100'."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return 0.0
    # Handle "100+100+100" strings
    try:
        return float(eval(s))  # Safe here: only digits and +
    except Exception:
        return 0.0


def read_excel(filepath: str) -> dict:
    """Read an Excel file and return contacts keyed by (NOM_upper, PRENOM_clean)."""
    contacts = OrderedDict()
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Feuil1"]

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        cells = list(row) + [None] * 5  # Pad to avoid index errors
        nom_raw = cells[0]
        prenom_raw = cells[1]
        email_raw = cells[2]
        adhesion = cells[3]
        cheque = cells[4]
        espece = cells[5]
        virement = cells[6]
        cb = cells[7]
        total_cot = cells[8]
        adresse = cells[9]

        if not nom_raw:
            continue

        nom = str(nom_raw).strip()
        prenom = clean_prenom(str(prenom_raw)) if prenom_raw else ""
        email = clean_email(email_raw)

        # Unique key: uppercase name + cleaned prenom
        key = (nom.upper(), prenom.upper())

        total = parse_total(total_cot)

        contacts[key] = {
            "nom": nom,
            "prenom": prenom,
            "email": email,
            "adhesion": 1 if adhesion == 1 else 0,
            "total": total,
            "adresse": str(adresse).strip() if adresse else "",
        }

    return contacts


def merge_contacts(contacts_2025: dict, contacts_2026: dict) -> list:
    """Merge 2025 and 2026 data. 2026 = base, 2025 = donation amounts."""
    merged = OrderedDict()

    # Start with 2026 as base (most current membership status)
    for key, data in contacts_2026.items():
        merged[key] = {
            "nom": data["nom"],
            "prenom": data["prenom"],
            "email": data["email"],
            "adhesion": data["adhesion"],
            "total_2026": data["total"],
            "total_2025": 0.0,
            "adresse": data["adresse"],
        }

    # Merge 2025 data (donation amounts + fill missing emails/addresses)
    for key, data in contacts_2025.items():
        if key in merged:
            merged[key]["total_2025"] = data["total"]
            # If 2026 is missing email but 2025 has one, use it
            if not merged[key]["email"] and data["email"]:
                merged[key]["email"] = data["email"]
            # Same for address
            if not merged[key]["adresse"] and data["adresse"]:
                merged[key]["adresse"] = data["adresse"]
        else:
            # Contact only in 2025, not in 2026
            merged[key] = {
                "nom": data["nom"],
                "prenom": data["prenom"],
                "email": data["email"],
                "adhesion": data["adhesion"],
                "total_2026": 0.0,
                "total_2025": data["total"],
                "adresse": data["adresse"],
            }

    return list(merged.values())


def determine_type(contact: dict) -> str:
    """Determine contact type based on membership and donation status."""
    is_adherent = contact["adhesion"] == 1
    is_donor = (contact["total_2025"] + contact["total_2026"]) > 0

    if is_adherent and is_donor:
        return "adhérent+donateur"
    elif is_adherent:
        return "adhérent"
    elif is_donor:
        return "donateur"
    else:
        return "contact"


def contact_to_brevo_row(contact: dict) -> dict:
    """Convert a contact to a Brevo CSV row."""
    addr = parse_address(contact["adresse"])
    total = contact["total_2025"] + contact["total_2026"]

    return {
        "EMAIL": contact["email"],
        "PRENOM": contact["prenom"],
        "NOM": contact["nom"],
        "ADRESSE": addr["full"],
        "CODE_POSTAL": addr["code_postal"],
        "COMMUNE": addr["commune"],
        "TYPE": determine_type(contact),
        "MONTANT": int(total) if total > 0 else "",
    }


def deduplicate_by_email(rows: list) -> tuple[list, list]:
    """Merge contacts sharing the same email into one Brevo contact.

    For couples/families: combines prénoms ("Jean Pierre & Nicole").
    Returns (deduplicated_rows, report_lines).
    """
    by_email = OrderedDict()
    for row in rows:
        email = row["EMAIL"]
        if email not in by_email:
            by_email[email] = []
        by_email[email].append(row)

    result = []
    report = []
    for email, group in by_email.items():
        if len(group) == 1:
            result.append(group[0])
            continue

        # Multiple contacts with same email — merge
        primary = group[0].copy()

        # Combine prénoms
        prenoms = [r["PRENOM"] for r in group if r["PRENOM"]]
        # Deduplicate prenoms (exact same name = true duplicate)
        seen = set()
        unique_prenoms = []
        for p in prenoms:
            if p.upper() not in seen:
                seen.add(p.upper())
                unique_prenoms.append(p)
        primary["PRENOM"] = " & ".join(unique_prenoms)

        # Combine noms if different
        noms = list(dict.fromkeys(r["NOM"] for r in group if r["NOM"]))
        primary["NOM"] = " / ".join(noms) if len(noms) > 1 else noms[0]

        # Best type: if any is adhérent, mark as adhérent
        types = [r["TYPE"] for r in group]
        has_adherent = any("adhérent" in t for t in types)
        has_donateur = any("donateur" in t for t in types)
        if has_adherent and has_donateur:
            primary["TYPE"] = "adhérent+donateur"
        elif has_adherent:
            primary["TYPE"] = "adhérent"
        elif has_donateur:
            primary["TYPE"] = "donateur"

        # Sum montants
        total = sum(int(r["MONTANT"]) for r in group if r["MONTANT"])
        primary["MONTANT"] = total if total > 0 else ""

        # Best address (take first non-empty)
        for r in group:
            if r["ADRESSE"]:
                primary["ADRESSE"] = r["ADRESSE"]
                primary["CODE_POSTAL"] = r["CODE_POSTAL"]
                primary["COMMUNE"] = r["COMMUNE"]
                break

        names_str = " + ".join(f"{r['PRENOM']} {r['NOM']}" for r in group)
        report.append(f"{email}: {names_str} → {primary['PRENOM']} {primary['NOM']}")
        result.append(primary)

    return result, report


def write_csv(filepath: str, rows: list, columns: list):
    """Write rows to a CSV file with BOM for Excel compatibility."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=columns, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)
    print(f"  {os.path.basename(filepath)}: {len(rows)} contacts")


def main():
    print("Lecture des fichiers Excel...")
    if not os.path.exists(FILE_2025):
        print(f"ERREUR: {FILE_2025} introuvable")
        sys.exit(1)
    if not os.path.exists(FILE_2026):
        print(f"ERREUR: {FILE_2026} introuvable")
        sys.exit(1)

    contacts_2025 = read_excel(FILE_2025)
    contacts_2026 = read_excel(FILE_2026)
    print(f"  2025: {len(contacts_2025)} contacts")
    print(f"  2026: {len(contacts_2026)} contacts")

    print("\nFusion des données...")
    all_contacts = merge_contacts(contacts_2025, contacts_2026)
    print(f"  Total après fusion: {len(all_contacts)} contacts")

    # Split by email presence
    with_email = [c for c in all_contacts if c["email"]]
    without_email = [c for c in all_contacts if not c["email"]]
    print(f"  Avec email: {len(with_email)}")
    print(f"  Sans email: {len(without_email)}")

    # Convert to Brevo rows
    brevo_rows_raw = [contact_to_brevo_row(c) for c in with_email]

    # Deduplicate by email (Brevo = 1 contact per email)
    brevo_rows, dupes_report = deduplicate_by_email(brevo_rows_raw)

    if dupes_report:
        print(f"\n  {len(dupes_report)} emails partagés (fusionnés) :")
        for line in dupes_report:
            print(f"    {line}")

    # Filter by type for list-specific CSVs
    adherent_rows = [r for r in brevo_rows if "adhérent" in r["TYPE"]]
    donateur_rows = [r for r in brevo_rows if "donateur" in r["TYPE"]]

    print(f"\nGénération des CSV dans {OUTPUT_DIR}/...")

    # All contacts with email -> "Tous les contacts ADESZ"
    write_csv(
        os.path.join(OUTPUT_DIR, "brevo-import-tous.csv"),
        brevo_rows,
        BREVO_COLUMNS,
    )

    # Adherents only -> "Adhérents ADESZ"
    write_csv(
        os.path.join(OUTPUT_DIR, "brevo-import-adherents.csv"),
        adherent_rows,
        BREVO_COLUMNS,
    )

    # Donors only -> "Donateurs ADESZ"
    write_csv(
        os.path.join(OUTPUT_DIR, "brevo-import-donateurs.csv"),
        donateur_rows,
        BREVO_COLUMNS,
    )

    # Contacts without email (for manual follow-up)
    no_email_rows = []
    for c in without_email:
        addr = parse_address(c["adresse"])
        no_email_rows.append({
            "NOM": c["nom"],
            "PRENOM": c["prenom"],
            "ADHESION": "oui" if c["adhesion"] == 1 else "non",
            "MONTANT_2025": int(c["total_2025"]) if c["total_2025"] > 0 else "",
            "MONTANT_2026": int(c["total_2026"]) if c["total_2026"] > 0 else "",
            "ADRESSE": addr["full"],
        })
    write_csv(
        os.path.join(OUTPUT_DIR, "contacts-sans-email.csv"),
        no_email_rows,
        ["NOM", "PRENOM", "ADHESION", "MONTANT_2025", "MONTANT_2026", "ADRESSE"],
    )

    # Print summary
    print("\n" + "=" * 50)
    print("RÉSUMÉ")
    print("=" * 50)
    print(f"Total contacts:       {len(all_contacts)}")
    print(f"Importables (email):  {len(with_email)}")
    print(f"  - Adhérents:        {len(adherent_rows)}")
    print(f"  - Donateurs:        {len(donateur_rows)}")
    print(f"Sans email:           {len(without_email)}")

    print("\n" + "=" * 50)
    print("PROCÉDURE D'IMPORT BREVO")
    print("=" * 50)
    print("""
1. Aller sur https://app.brevo.com/contacts/import

2. Import n°1 — Tous les contacts :
   - Fichier : brevo-import-tous.csv
   - Liste : "Tous les contacts ADESZ"
   - Mapper les colonnes (EMAIL, PRENOM, NOM, etc.)

3. Import n°2 — Adhérents :
   - Fichier : brevo-import-adherents.csv
   - Liste : "Adhérents ADESZ"

4. Import n°3 — Donateurs :
   - Fichier : brevo-import-donateurs.csv
   - Liste : "Donateurs ADESZ"

Note : les 3 imports se font dans cet ordre. Les contacts
déjà existants seront mis à jour (pas de doublons).
""")


if __name__ == "__main__":
    main()
